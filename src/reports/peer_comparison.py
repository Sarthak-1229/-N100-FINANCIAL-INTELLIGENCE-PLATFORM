"""
src/reports/peer_comparison.py

Peer Comparison Report Generator - creates Excel reports for each peer group
with percentile rankings and color-coding.
"""
import os
import sqlite3
import pandas as pd

DB_PATH = os.getenv("DB_PATH", "db/nifty100.db")
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "output")


def generate_peer_comparison_report(year: str = "2024-03", db_path: str = DB_PATH):
    """
    Generate peer comparison Excel report with one sheet per peer group.

    Args:
        year: Fiscal year
        db_path: Database path

    Returns:
        Path to generated Excel file
    """
    conn = sqlite3.connect(db_path)

    # Get peer groups with company info and is_benchmark flag
    peer_groups_query = """
        SELECT pg.peer_group_name, pg.company_id, pg.is_benchmark,
               c.company_name, s.broad_sector, s.sub_sector
        FROM peer_groups pg
        JOIN companies c ON pg.company_id = c.id
        LEFT JOIN sectors s ON pg.company_id = s.company_id
    """
    peer_groups_df = pd.read_sql(peer_groups_query, conn)

    # Get financial data for the year
    financial_query = """
        SELECT fr.* FROM financial_ratios fr WHERE fr.year = ?
    """
    financial_df = pd.read_sql(financial_query, conn, params=[year])

    # Get peer percentiles for the year
    percentiles_query = """
        SELECT pp.peer_group_name, pp.company_id, pp.metric,
               pp.value, pp.percentile_rank
        FROM peer_percentiles pp
        WHERE pp.year = ?
    """
    percentiles_df = pd.read_sql(percentiles_query, conn, params=[year])

    conn.close()

    # Merge all data
    merged = peer_groups_df.merge(
        financial_df, on='company_id', how='left'
    ).merge(
        percentiles_df, on=['peer_group_name', 'company_id', 'metric'], how='left'
    )

    # Prepare metrics for percentile rank columns
    metrics = [
        'ROE', 'ROCE', 'NPM', 'DE', 'FCF',
        'PAT_CAGR_5YR', 'REVENUE_CAGR_5YR', 'EPS_CAGR_5YR',
        'ICR', 'ASSET_TURNOVER'
    ]

    # Pivot to get one row per company with columns for each metric's value and percentile
    value_pivot = merged.pivot_table(
        index=['peer_group_name', 'company_id', 'company_name', 'is_benchmark', 'broad_sector', 'sub_sector'],
        columns='metric',
        values='value',
        aggfunc='first'
    ).reset_index()

    percentile_pivot = merged.pivot_table(
        index=['peer_group_name', 'company_id', 'company_name', 'is_benchmark', 'broad_sector', 'sub_sector'],
        columns='metric',
        values='percentile_rank',
        aggfunc='first'
    ).reset_index()

    # Rename percentile columns
    percentile_pivot = percentile_pivot.rename(columns={
        metric: f'pct_rank_{metric}' for metric in metrics
    })

    # Merge value and percentile data
    result_df = value_pivot.merge(
        percentile_pivot,
        on=['peer_group_name', 'company_id', 'company_name', 'is_benchmark', 'broad_sector', 'sub_sector']
    )

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")

    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Define colors
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        gold_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Process each peer group
        for peer_group in sorted(result_df['peer_group_name'].unique()):
            group_data = result_df[result_df['peer_group_name'] == peer_group].copy()

            if group_data.empty:
                continue

            # Add a summary row with medians
            summary_row = {}
            for col in group_data.columns:
                if col in ['company_id', 'company_name', 'is_benchmark', 'broad_sector', 'sub_sector']:
                    summary_row[col] = 'MEDIAN'
                elif col.startswith('pct_rank_'):
                    # Median of percentile ranks
                    median_val = group_data[col].median()
                    summary_row[col] = median_val
                else:
                    # Median of values
                    median_val = group_data[col].median()
                    summary_row[col] = median_val

            # Add summary row
            summary_df = pd.DataFrame([summary_row])
            group_with_summary = pd.concat([group_data, summary_df], ignore_index=True)

            # Write to sheet (limit sheet name to 31 chars)
            sheet_name = peer_group[:31]
            group_with_summary.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get worksheet for formatting
            ws = writer.sheets[sheet_name]

            # Header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment

            # Color percentile columns
            percentile_cols = [col for col in group_data.columns if col.startswith('pct_rank_')]
            for col_idx, col_name in enumerate(ws[1], 1):  # Row 1 is header
                if col_name.value in percentile_cols:
                    # Apply color coding to each cell in this column
                    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            val = float(cell.value) if cell.value is not None else 0
                            if val >= 0.75:
                                cell.fill = green_fill
                            elif val >= 0.25:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill
                        except (ValueError, TypeError):
                            pass

            # Highlight benchmark row (gold background)
            for row_idx in range(2, ws.max_row + 1):  # Skip header
                is_benchmark_cell = ws.cell(row=row_idx, column=ws.min_column +
                                          list(group_with_summary.columns).index('is_benchmark'))
                try:
                    if is_benchmark_cell.value == 1 or str(is_benchmark_cell.value).lower() == 'true':
                        # Highlight entire row
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = gold_fill
                except:
                    pass

            # Add summary row label (bold)
            summary_row_idx = ws.max_row  # Last row is summary
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=summary_row_idx, column=col_idx)
                cell.font = Font(bold=True)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    print(f"Generated peer comparison report: {output_path}")
    return output_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate peer comparison report")
    parser.add_argument("--year", type=str, default="2024-03", help="Fiscal year")
    parser.add_argument("--db-path", type=str, default=DB_PATH, help="Database path")

    args = parser.parse_args()

    generate_peer_comparison_report(year=args.year, db_path=args.db_path)