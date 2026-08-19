"""
src/screener/engine.py

Financial Screener Engine - applies preset or custom filters to financial_ratios.
Supports 6 preset screeners, custom thresholds, composite quality score,
and exports to color-coded Excel.
"""
import os
import sqlite3
from typing import Optional

import pandas as pd
import yaml


DB_PATH = os.getenv("DB_PATH", "db/nifty100.db")
CONFIG_PATH = os.getenv("SCREENER_CONFIG", "config/screener_config.yaml")
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "output")


def load_config():
    """Load screener configuration from YAML."""
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


def get_financials_companies(conn: sqlite3.Connection) -> set:
    """Get set of company IDs in Financials broad_sector."""
    cur = conn.execute("""
        SELECT c.id FROM companies c
        JOIN sectors s ON c.id = s.company_id
        WHERE s.broad_sector = 'Financials'
    """)
    return {row[0] for row in cur.fetchall()}


def load_screener_data(year: str = "2024-03") -> pd.DataFrame:
    """
    Load latest financial_ratios + market_cap + P&L for screener.
    Joins on company_id and year (market_cap.year is integer, so match year prefix).
    """
    conn = sqlite3.connect(DB_PATH)

    # Extract year prefix from the year parameter (e.g., '2024-03' -> '2024')
    year_prefix = year[:4]

    query = """
        SELECT
            fr.*,
            mc.pe_ratio,
            mc.pb_ratio,
            mc.dividend_yield_pct,
            mc.market_cap_crore,
            mc.enterprise_value_crore,
            mc.ev_ebitda,
            pl.net_profit,
            pl.sales,
            pl.operating_profit,
            pl.other_income,
            pl.interest,
            pl.depreciation,
            pl.eps,
            pl.dividend_payout,
            c.company_name,
            s.broad_sector,
            s.sub_sector
        FROM financial_ratios fr
        JOIN market_cap mc ON fr.company_id = mc.company_id
            AND CAST(mc.year AS TEXT) = ?
        JOIN profitandloss pl ON fr.company_id = pl.company_id
            AND fr.year = pl.year
        JOIN companies c ON fr.company_id = c.id
        LEFT JOIN sectors s ON fr.company_id = s.company_id
        WHERE fr.year = ?
    """

    df = pd.read_sql(query, conn, params=[year_prefix, year])
    conn.close()
    return df


def apply_filters(df: pd.DataFrame, filters: dict, financials_ids: set) -> pd.DataFrame:
    """
    Apply filter dictionary to DataFrame.

    Special handling:
    - de_max: skip Financials companies
    - icr_min: treat "Debt Free" (None interest_coverage) as infinity
    """
    result = df.copy()
    original_count = len(result)

    for filter_name, threshold in filters.items():
        if threshold is None:
            continue

        col = filter_name.replace('_min', '').replace('_max', '')
        is_min = filter_name.endswith('_min')
        is_max = filter_name.endswith('_max')

        # Map filter names to actual columns
        col_map = {
            'roe': 'return_on_equity_pct',
            'de': 'debt_to_equity',
            'fcf': 'free_cash_flow_cr',
            'revenue_cagr_5yr': 'revenue_cagr_5yr',
            'pat_cagr_5yr': 'pat_cagr_5yr',
            'opm': 'operating_profit_margin_pct',
            'pe': 'pe_ratio',
            'pb': 'pb_ratio',
            'div_yield': 'dividend_yield_pct',
            'icr': 'interest_coverage',
            'market_cap': 'market_cap_crore',
            'net_profit': 'net_profit',
            'eps_cagr': 'eps_cagr_5yr',
            'asset_turnover': 'asset_turnover',
            'sales': 'sales',
            'dividend_payout': 'dividend_payout_pct',
            'revenue_cagr_3yr': 'revenue_cagr_3yr',
            'fcf_latest_positive': 'fcf_latest_positive',
            'de_declining_yoy': 'de_declining_yoy',
        }

        actual_col = col_map.get(col, col)

        if actual_col not in result.columns:
            print(f"Warning: Column {actual_col} not found for filter {filter_name}")
            continue

        if is_min:
            mask = result[actual_col] >= threshold
        else:  # is_max
            mask = result[actual_col] <= threshold

        # Special: D/E max filter - skip Financials companies
        if filter_name == 'de_max':
            financials_mask = result['company_id'].isin(financials_ids)
            mask = mask | financials_mask  # Financials always pass D/E filter

        # Special: ICR min filter - Debt Free (None) = infinity, always passes
        if filter_name == 'icr_min':
            debt_free_mask = result[actual_col].isna()  # Debt Free companies have None
            mask = mask | debt_free_mask

        result = result[mask].copy()

        filtered_out = original_count - len(result)
        if filtered_out > 0:
            print(f"  {filter_name} ({threshold}): filtered out {filtered_out} companies")
            original_count = len(result)

    return result


def compute_composite_score(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute composite quality score (0-100) with sector-relative normalization.

    Weights:
    35% Profitability: ROE (15%), ROCE (10%), NPM (10%)
    30% Cash Quality: FCF CAGR 5yr (15%), CFO/PAT avg 5yr (10%), FCF positive (5%)
    20% Growth: Rev CAGR 5yr (10%), PAT CAGR 5yr (10%)
    15% Leverage: D/E score (10%), ICR score (5%)
    """
    result = df.copy()

    # Component scores (0-100 each)
    # We'll use P10/P90 winsorization for normalization

    def winsorize_normalize(series: pd.Series, invert: bool = False) -> pd.Series:
        """Normalize to 0-100 using P10/P90 winsorization."""
        if series.isna().all():
            return pd.Series(50.0, index=series.index)  # neutral score

        # Winsorize at P10/P90
        p10 = series.quantile(0.10)
        p90 = series.quantile(0.90)

        if p10 == p90:
            return pd.Series(50.0, index=series.index)

        clipped = series.clip(p10, p90)
        normalized = (clipped - p10) / (p90 - p10) * 100

        if invert:
            normalized = 100 - normalized

        return normalized.fillna(50.0)

    # We need sector-specific P10/P90
    sectors = result['broad_sector'].fillna('Unknown').unique()

    # Initialize component scores
    component_scores = pd.DataFrame(index=result.index)

    # Profitability (35%)
    component_scores['roe_score'] = winsorize_normalize(result['return_on_equity_pct'])
    component_scores['roce_score'] = winsorize_normalize(result['return_on_capital_employed_pct'])
    component_scores['npm_score'] = winsorize_normalize(result['net_profit_margin_pct'])

    # Cash Quality (30%)
    # FCF CAGR 5yr
    component_scores['fcf_cagr_score'] = winsorize_normalize(result['pat_cagr_5yr'])  # Using PAT CAGR as proxy
    # CFO/PAT ratio - use capital_allocation.csv data if available
    cfo_pat_score = result.get('cfo_pat_ratio', pd.Series(index=result.index, dtype=float))
    if cfo_pat_score.isna().all():
        # Use free_cash_flow_cr / net_profit as proxy
        with pd.option_context('mode.chained_assignment', None):
            cfo_pat = result['free_cash_flow_cr'] / result['net_profit']
        component_scores['cfo_pat_score'] = winsorize_normalize(cfo_pat)
    else:
        component_scores['cfo_pat_score'] = winsorize_normalize(cfo_pat_score)
    # FCF positive flag
    component_scores['fcf_positive_score'] = (result['free_cash_flow_cr'] > 0).astype(float) * 100

    # Growth (20%)
    component_scores['rev_cagr_score'] = winsorize_normalize(result['revenue_cagr_5yr'])
    component_scores['pat_cagr_score'] = winsorize_normalize(result['pat_cagr_5yr'])

    # Leverage (15%) - D/E and ICR inverted (lower D/E = better, higher ICR = better)
    component_scores['de_score'] = winsorize_normalize(result['debt_to_equity'], invert=True)
    component_scores['icr_score'] = winsorize_normalize(result['interest_coverage'])

    # Weighted composite
    weights = {
        'roe_score': 0.15,
        'roce_score': 0.10,
        'npm_score': 0.10,
        'fcf_cagr_score': 0.15,
        'cfo_pat_score': 0.10,
        'fcf_positive_score': 0.05,
        'rev_cagr_score': 0.10,
        'pat_cagr_score': 0.10,
        'de_score': 0.10,
        'icr_score': 0.05,
    }

    result['composite_score'] = 0.0
    for col, weight in weights.items():
        if col in component_scores.columns:
            result['composite_score'] += component_scores[col] * weight

    # Sector-relative normalization
    result['sector_relative_score'] = result.groupby('broad_sector')['composite_score'].transform(
        lambda x: (x - x.min()) / (x.max() - x.min()) * 100 if x.max() != x.min() else 50
    ).fillna(50)

    # Final score = average of absolute and sector-relative
    result['final_score'] = (result['composite_score'] + result['sector_relative_score']) / 2

    return result


def run_screener(
    db_path: str = DB_PATH,
    preset: Optional[str] = None,
    custom_filters: Optional[dict] = None,
    year: str = "2024-03"
) -> pd.DataFrame:
    """
    Main entry point for running the screener.

    Args:
        db_path: Path to SQLite database
        preset: Name of preset from config (e.g., 'quality_compounder')
        custom_filters: Dict of filter_name -> threshold
        year: Fiscal year to screen (default latest available)

    Returns:
        DataFrame with filtered companies, sorted by final_score DESC
    """
    config = load_config()

    # Load data
    print(f"Loading data for {year}...")
    df = load_screener_data(year)
    print(f"  Loaded {len(df)} companies")

    # Get Financials companies
    conn = sqlite3.connect(db_path)
    financials_ids = get_financials_companies(conn)
    conn.close()

    # Determine filters
    if preset:
        if preset not in config['presets']:
            raise ValueError(f"Unknown preset: {preset}. Available: {list(config['presets'].keys())}")
        filters = config['presets'][preset].copy()
        # Remove description
        filters.pop('description', None)
        print(f"Running preset: {preset}")
    elif custom_filters:
        filters = custom_filters
        print("Running custom filters")
    else:
        filters = {}
        print("Running without filters (all companies)")

    # Apply filters
    df_filtered = apply_filters(df, filters, financials_ids)
    print(f"After filters: {len(df_filtered)} companies")

    # Compute composite score
    df_scored = compute_composite_score(df_filtered)

    # Sort by final_score DESC
    df_scored = df_scored.sort_values('final_score', ascending=False).reset_index(drop=True)

    return df_scored


def run_all_presets(year: str = "2024-03") -> dict:
    """Run all 6 presets and return dict of DataFrames."""
    config = load_config()
    results = {}

    for preset_name in config['presets'].keys():
        print(f"\n{'='*50}")
        print(f"Running preset: {preset_name}")
        print(f"{'='*50}")
        df = run_screener(preset=preset_name, year=year)
        results[preset_name] = df
        print(f"  Result: {len(df)} companies")

    return results


def export_screener_output(results: dict, output_path: str = None):
    """Export all preset results to color-coded Excel."""
    if output_path is None:
        output_path = os.path.join(OUTPUT_DIR, "screener_output.xlsx")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        for preset_name, df in results.items():
            if df.empty:
                continue

            # Select display columns
            display_cols = [
                'company_id', 'company_name', 'broad_sector', 'sub_sector',
                'final_score', 'composite_score', 'sector_relative_score',
                'return_on_equity_pct', 'return_on_capital_employed_pct',
                'net_profit_margin_pct', 'operating_profit_margin_pct',
                'debt_to_equity', 'interest_coverage',
                'free_cash_flow_cr', 'revenue_cagr_5yr', 'pat_cagr_5yr',
                'eps_cagr_5yr', 'asset_turnover', 'pe_ratio', 'pb_ratio',
                'dividend_yield_pct', 'market_cap_crore', 'net_profit',
                'capital_allocation_pattern'
            ]

            # Only keep columns that exist
            display_cols = [c for c in display_cols if c in df.columns]
            df_display = df[display_cols].copy()

            # Write to sheet
            df_display.to_excel(writer, sheet_name=preset_name[:31], index=False)

            # Get worksheet for formatting
            ws = writer.sheets[preset_name[:31]]

            # Apply color coding for filter columns
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

            # Auto-width columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    print(f"Exported screener output to {output_path}")


def export_peer_comparison(peer_groups: dict, output_path: str = None):
    """Export peer comparison report to Excel."""
    if output_path is None:
        output_path = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        from openpyxl.styles import PatternFill, Font, Alignment

        for group_name, df in peer_groups.items():
            if df.empty:
                continue

            # Sort by percentile rank for key metric
            pct_cols = [c for c in df.columns if c.startswith('pct_rank_')]
            if pct_cols:
                df = df.sort_values(pct_cols[0], ascending=False)

            sheet_name = group_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            # Color-code percentile columns
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            gold_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

            # Highlight benchmark row
            for row_idx, row in df.iterrows():
                if row.get('is_benchmark', False):
                    for cell in ws[row_idx + 2]:  # +2 for header + 1-indexed
                        cell.fill = gold_fill

            # Color percentile columns
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name.startswith('pct_rank_'):
                    for row_idx in range(2, len(df) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            val = float(cell.value) if cell.value is not None else 0
                            if val >= 0.75:
                                cell.fill = green_fill
                            elif val >= 0.25:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill
                        except (ValueError, TypeError):
                            pass

            # Auto-width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width

    print(f"Exported peer comparison to {output_path}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Financial Screener")
    parser.add_argument("--preset", type=str, help="Preset name to run")
    parser.add_argument("--all-presets", action="store_true", help="Run all presets")
    parser.add_argument("--year", type=str, default="2024-03", help="Fiscal year")
    parser.add_argument("--output", type=str, help="Output Excel path")

    args = parser.parse_args()

    if args.all_presets:
        results = run_all_presets(year=args.year)
        export_screener_output(results, args.output)
    elif args.preset:
        df = run_screener(preset=args.preset, year=args.year)
        print(f"\nTop 10 results:")
        print(df[['company_id', 'company_name', 'final_score', 'return_on_equity_pct', 'debt_to_equity']].head(10).to_string(index=False))
        if args.output:
            export_screener_output({args.preset: df}, args.output)
    else:
        print("Please specify --preset or --all-presets")